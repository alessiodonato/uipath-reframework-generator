#!/usr/bin/env python3
"""
Config.xlsx Manager for UiPath ReFramework Projects.

Provides CLI for managing Config.xlsx files:
- Add/update keys
- List all keys
- Validate keys against XAML usage
- Cross-reference detection

Usage:
    # List all keys
    python config_manager.py list ./MyProject

    # Add a key
    python config_manager.py add ./MyProject --sheet Settings --key AppUrl --value "https://..." --desc "App URL"

    # Validate keys (find unused and missing)
    python config_manager.py validate ./MyProject

    # Export as JSON
    python config_manager.py export ./MyProject --json
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG.XLSX OPERATIONS
# ═══════════════════════════════════════════════════════════════════════════════

def load_config(project_dir: str) -> Dict[str, List[Dict]]:
    """
    Load Config.xlsx and return all sheets as dicts.

    Args:
        project_dir: Path to project directory

    Returns:
        Dict of {sheet_name: [{Name, Value, Description}, ...]}
    """
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    config_path = Path(project_dir) / "Data" / "Config.xlsx"
    if not config_path.exists():
        print(f"Error: Config.xlsx not found at {config_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(config_path)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header

        sheet_data = []
        for row in rows:
            if row[0]:  # Has a name
                sheet_data.append({
                    "Name": str(row[0]) if row[0] else "",
                    "Value": str(row[1]) if row[1] else "",
                    "Description": str(row[2]) if len(row) > 2 and row[2] else "",
                })
        result[sheet_name] = sheet_data

    return result


def save_config(project_dir: str, data: Dict[str, List[Dict]]):
    """
    Save config data back to Config.xlsx.

    Args:
        project_dir: Path to project directory
        data: Dict of {sheet_name: [{Name, Value, Description}, ...]}
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    config_path = Path(project_dir) / "Data" / "Config.xlsx"

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    first_sheet = True
    for sheet_name, rows in data.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        # Headers
        for col, header in enumerate(["Name", "Value", "Description"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col, key in enumerate(["Name", "Value", "Description"], 1):
                cell = ws.cell(row=row_idx, column=col, value=row_data.get(key, ""))
                cell.border = thin_border
                if fill:
                    cell.fill = fill

        # Column widths
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 50
        ws.freeze_panes = "A2"

    wb.save(config_path)


def add_key(
    project_dir: str,
    sheet: str,
    key: str,
    value: str,
    description: str = ""
) -> bool:
    """
    Add or update a key in Config.xlsx.

    Args:
        project_dir: Path to project directory
        sheet: Sheet name (Settings, Constants, Assets)
        key: Key name
        value: Key value
        description: Optional description

    Returns:
        True if key was added/updated
    """
    data = load_config(project_dir)

    if sheet not in data:
        data[sheet] = []

    # Check if key exists
    for row in data[sheet]:
        if row["Name"] == key:
            row["Value"] = value
            if description:
                row["Description"] = description
            save_config(project_dir, data)
            print(f"Updated: {sheet}/{key} = {value}")
            return True

    # Add new key
    data[sheet].append({
        "Name": key,
        "Value": value,
        "Description": description,
    })
    save_config(project_dir, data)
    print(f"Added: {sheet}/{key} = {value}")
    return True


def remove_key(project_dir: str, sheet: str, key: str) -> bool:
    """
    Remove a key from Config.xlsx.

    Args:
        project_dir: Path to project directory
        sheet: Sheet name
        key: Key name

    Returns:
        True if key was removed
    """
    data = load_config(project_dir)

    if sheet not in data:
        print(f"Sheet '{sheet}' not found", file=sys.stderr)
        return False

    original_len = len(data[sheet])
    data[sheet] = [row for row in data[sheet] if row["Name"] != key]

    if len(data[sheet]) < original_len:
        save_config(project_dir, data)
        print(f"Removed: {sheet}/{key}")
        return True
    else:
        print(f"Key '{key}' not found in {sheet}", file=sys.stderr)
        return False


def list_keys(project_dir: str, as_json: bool = False) -> Dict[str, List[Dict]]:
    """
    List all keys in Config.xlsx.

    Args:
        project_dir: Path to project directory
        as_json: Return as JSON string

    Returns:
        Dict of all config data
    """
    data = load_config(project_dir)

    if as_json:
        print(json.dumps(data, indent=2))
    else:
        for sheet, rows in data.items():
            print(f"\n{sheet}:")
            print("-" * 60)
            for row in rows:
                desc = f" ({row['Description'][:40]})" if row['Description'] else ""
                print(f"  {row['Name']}: {row['Value']}{desc}")

    return data


# ═══════════════════════════════════════════════════════════════════════════════
# VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

def extract_config_references(xaml_content: str) -> Set[str]:
    """
    Extract all Config() references from XAML content.

    Args:
        xaml_content: XAML file content

    Returns:
        Set of config key names
    """
    refs = set()

    # Pattern: Config("key") or Config('key') or in_Config("key")
    patterns = [
        r'Config\("([^"]+)"\)',
        r"Config\('([^']+)'\)",
        r'in_Config\("([^"]+)"\)',
        r"in_Config\('([^']+)'\)",
        r'Config\(&quot;([^&]+)&quot;\)',
    ]

    for pattern in patterns:
        refs.update(re.findall(pattern, xaml_content))

    return refs


def get_all_config_references(project_dir: str) -> Dict[str, Set[str]]:
    """
    Get all Config() references from all XAML files.

    Args:
        project_dir: Path to project directory

    Returns:
        Dict of {xaml_file: {keys...}}
    """
    project_path = Path(project_dir)
    refs = {}

    for xaml_file in project_path.rglob("*.xaml"):
        content = xaml_file.read_text(encoding="utf-8")
        file_refs = extract_config_references(content)
        if file_refs:
            rel_path = str(xaml_file.relative_to(project_path))
            refs[rel_path] = file_refs

    return refs


def validate_config(project_dir: str) -> Tuple[Set[str], Set[str]]:
    """
    Validate Config.xlsx against XAML usage.

    Args:
        project_dir: Path to project directory

    Returns:
        Tuple of (missing_keys, unused_keys)
    """
    # Get all defined keys
    data = load_config(project_dir)
    defined_keys = set()
    for rows in data.values():
        for row in rows:
            defined_keys.add(row["Name"])

    # Get all referenced keys
    refs = get_all_config_references(project_dir)
    referenced_keys = set()
    for file_refs in refs.values():
        referenced_keys.update(file_refs)

    # Find mismatches
    missing_keys = referenced_keys - defined_keys
    unused_keys = defined_keys - referenced_keys

    # Report
    print("\nConfig.xlsx Validation Report")
    print("=" * 60)

    if missing_keys:
        print(f"\n{len(missing_keys)} MISSING keys (used in XAML but not in Config.xlsx):")
        for key in sorted(missing_keys):
            # Find which files use this key
            files = [f for f, keys in refs.items() if key in keys]
            print(f"  - {key}")
            for f in files[:3]:
                print(f"      Used in: {f}")
            if len(files) > 3:
                print(f"      ... and {len(files) - 3} more files")
    else:
        print("\nNo missing keys.")

    if unused_keys:
        print(f"\n{len(unused_keys)} UNUSED keys (in Config.xlsx but not referenced in XAML):")
        for key in sorted(unused_keys):
            print(f"  - {key}")
    else:
        print("\nNo unused keys.")

    print("\n" + "=" * 60)
    print(f"Total defined: {len(defined_keys)}  |  "
          f"Total referenced: {len(referenced_keys)}  |  "
          f"Missing: {len(missing_keys)}  |  "
          f"Unused: {len(unused_keys)}")

    return missing_keys, unused_keys


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Manage Config.xlsx for UiPath ReFramework projects",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    subparsers = parser.add_subparsers(dest="command", help="Command")

    # list command
    list_parser = subparsers.add_parser("list", help="List all config keys")
    list_parser.add_argument("project", help="Path to project directory")
    list_parser.add_argument("--json", "-j", action="store_true", help="Output as JSON")

    # add command
    add_parser = subparsers.add_parser("add", help="Add or update a config key")
    add_parser.add_argument("project", help="Path to project directory")
    add_parser.add_argument("--sheet", "-s", required=True,
                           choices=["Settings", "Constants", "Assets"],
                           help="Sheet name")
    add_parser.add_argument("--key", "-k", required=True, help="Key name")
    add_parser.add_argument("--value", "-v", required=True, help="Key value")
    add_parser.add_argument("--desc", "-d", default="", help="Description")

    # remove command
    remove_parser = subparsers.add_parser("remove", help="Remove a config key")
    remove_parser.add_argument("project", help="Path to project directory")
    remove_parser.add_argument("--sheet", "-s", required=True, help="Sheet name")
    remove_parser.add_argument("--key", "-k", required=True, help="Key name")

    # validate command
    validate_parser = subparsers.add_parser("validate",
                                            help="Validate config against XAML usage")
    validate_parser.add_argument("project", help="Path to project directory")

    # export command
    export_parser = subparsers.add_parser("export", help="Export config as JSON")
    export_parser.add_argument("project", help="Path to project directory")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    if args.command == "list":
        list_keys(args.project, as_json=args.json)
    elif args.command == "add":
        add_key(args.project, args.sheet, args.key, args.value, args.desc)
    elif args.command == "remove":
        remove_key(args.project, args.sheet, args.key)
    elif args.command == "validate":
        missing, unused = validate_config(args.project)
        if missing:
            sys.exit(1)
    elif args.command == "export":
        list_keys(args.project, as_json=True)


if __name__ == "__main__":
    main()
